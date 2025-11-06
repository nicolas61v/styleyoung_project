"""
Implementación concreta de ReporteInterface para generar reportes en Excel
"""
from typing import List, Dict, Any
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
from .reporte_interface import ReporteInterface


class ReporteExcel(ReporteInterface):
    """
    Generador de reportes en formato Excel

    Implementación concreta que genera reportes utilizando openpyxl.
    Esta clase puede ser utilizada sin modificar el código cliente gracias
    a la inversión de dependencias.
    """

    def generar_reporte(self, titulo: str, datos: List[Dict[str, Any]], columnas: List[str]) -> HttpResponse:
        """
        Genera un reporte en formato Excel (XLSX)

        Args:
            titulo: Título del reporte
            datos: Lista de diccionarios con los datos
            columnas: Lista de columnas a incluir en el reporte

        Returns:
            HttpResponse con el archivo Excel generado
        """
        # Crear libro de trabajo
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        title_font = Font(bold=True, size=16)
        title_alignment = Alignment(horizontal="center", vertical="center")

        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título del reporte
        ws.merge_cells(f'A1:{get_column_letter(len(columnas))}1')
        title_cell = ws['A1']
        title_cell.value = titulo
        title_cell.font = title_font
        title_cell.alignment = title_alignment

        # Fecha de generación
        ws.merge_cells(f'A2:{get_column_letter(len(columnas))}2')
        fecha_cell = ws['A2']
        fecha_cell.value = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        fecha_cell.alignment = Alignment(horizontal="center")

        # Espacio en blanco
        # Fila 3 vacía

        # Encabezados (fila 4)
        for col_num, columna in enumerate(columnas, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = columna
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Datos (desde fila 5)
        for row_num, item in enumerate(datos, 5):
            for col_num, columna in enumerate(columnas, 1):
                cell = ws.cell(row=row_num, column=col_num)
                valor = item.get(columna, '')

                # Formatear valores según tipo
                if isinstance(valor, (int, float)):
                    cell.value = valor
                    if 'precio' in columna.lower() or 'total' in columna.lower():
                        cell.number_format = '"$"#,##0.00'
                else:
                    cell.value = str(valor)

                cell.border = border_style
                cell.alignment = Alignment(horizontal="left", vertical="center")

                # Alternar colores de fila
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="f8fafc", end_color="f8fafc", fill_type="solid")

        # Ajustar ancho de columnas
        for col_num, columna in enumerate(columnas, 1):
            column_letter = get_column_letter(col_num)
            # Calcular ancho basado en el contenido
            max_length = len(columna)
            for row_num in range(5, len(datos) + 5):
                cell_value = ws.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

        # Pie de página
        footer_row = len(datos) + 6
        ws.merge_cells(f'A{footer_row}:{get_column_letter(len(columnas))}{footer_row}')
        footer_cell = ws.cell(row=footer_row, column=1)
        footer_cell.value = "StyleYoung - Tienda Virtual de Ropa"
        footer_cell.font = Font(italic=True)
        footer_cell.alignment = Alignment(horizontal="center")

        # Total de registros
        total_row = footer_row + 1
        ws.merge_cells(f'A{total_row}:{get_column_letter(len(columnas))}{total_row}')
        total_cell = ws.cell(row=total_row, column=1)
        total_cell.value = f"Total de registros: {len(datos)}"
        total_cell.alignment = Alignment(horizontal="center")

        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(
            buffer.getvalue(),
            content_type=self.get_content_type()
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.{self.get_extension()}"'

        buffer.close()
        return response

    def get_content_type(self) -> str:
        """Retorna el content type para Excel"""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def get_extension(self) -> str:
        """Retorna la extensión de archivo para Excel"""
        return 'xlsx'
